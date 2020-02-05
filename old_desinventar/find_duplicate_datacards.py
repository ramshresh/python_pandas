"""
https://thispointer.com/pandas-find-duplicate-rows-in-a-dataframe-based-on-all-or-selected-columns-using-dataframe-duplicated-in-python/
"""
import pandas as pd

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    Usage examples:

append_df_to_excel('d:/temp/test.xlsx', df)

append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', index=False)

append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', index=False, startrow=25)

    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

    
df = pd.read_excel(r"DesInventar2016.xlsx", "DesInventar2016")

duplicateRowsDF = df[df.duplicated()]


print("Duplicate Rows except first occurrence based on all columns are :")
print(duplicateRowsDF)

duplicateRowsDF.to_excel(r"DesInventar2016-duplicates_all_columns.xlsx")


duplicateRowsDF_SN_CENTER = df[df.duplicated(['Serial_No', 'CENTER'])]
 
print("Duplicate Rows based on Serial_No and CENTER columns are:", duplicateRowsDF_SN_CENTER, sep='\n')

duplicateRowsDF_SN_CENTER.to_excel(r"DesInventar2016-duplicates_all_Serial_No_and_CENTER.xlsx")

duplicateRowsDF_SN = df[df.duplicated(['Serial_No'])]
 
print("Duplicate Rows based on Serial_No are:", duplicateRowsDF_SN, sep='\n')

duplicateRowsDF_SN.to_excel(r"DesInventar2016-duplicates_all_Serial_No.xlsx")



duplicateRowsDF_SN_YMDE = df[df.duplicated(['Serial_No','Year','Month', 'Day', 'Event'])]
 
print("Duplicate Rows based on SN, Year, Month, Day, Event are:", duplicateRowsDF_SN_YMDE, sep='\n')

duplicateRowsDF_SN_YMDE.to_excel(r"DesInventar2016-duplicates_all_SN_YMDE.xlsx")



data_cols=['Serial_No','Dead_Peopl','Missing_Pe', 'Injured_Pe', 'Victims',
                                            'Affected_P','Destroyed','Affected_H','Evacuated','Affected_R',
                                            'Farming_an','Livestock','Education','Relocated', 'Medical_Ce',
                                            'Losses_Val', 'Losses__US', 'Magnitude','Other_Loss'
                                            ]
duplicateRowsDF_SN_Data = df[df.duplicated(data_cols)]
 
print("Duplicate Rows based on SN, "+"#".join(data_cols), duplicateRowsDF_SN_Data, sep='\n')

duplicateRowsDF_SN_Data.to_excel(r"DesInventar2016-duplicates_all_SN_Data.xlsx")




print (duplicateRowsDF['Affected_R'].sum())

print (duplicateRowsDF_SN_CENTER['Affected_R'].sum())

print (duplicateRowsDF_SN['Affected_R'].sum())
print (duplicateRowsDF_SN_YMDE['Affected_R'].sum())

print (duplicateRowsDF_SN_Data['Affected_R'].sum())


append_df_to_excel(r"DesInventar2016-duplicates.xlsx",duplicateRowsDF, sheet_name="dup_all_Data", startrow=0 )
append_df_to_excel(r"DesInventar2016-duplicates.xlsx",duplicateRowsDF_SN, sheet_name="dup_SN",startrow=0 )
append_df_to_excel(r"DesInventar2016-duplicates.xlsx",duplicateRowsDF_SN_CENTER, sheet_name="dup_SN_CENTER",startrow=0 )
append_df_to_excel(r"DesInventar2016-duplicates.xlsx",duplicateRowsDF_SN_YMDE, sheet_name="dup_SN_YMDE",startrow=0 )
append_df_to_excel(r"DesInventar2016-duplicates.xlsx",duplicateRowsDF_SN_Data, sheet_name="dup_SN_Data",startrow=0 )


